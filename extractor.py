import re
import tkinter as tk
from openpyxl import load_workbook


patterns = {
    "last_first_name": re.compile(r"(?P<last_name>(?:\w\s?)*), (?P<first_name>[A-Z\u00C0-\u00DD][a-z\u00E0-\u00FF]+(?:\s?\w)*)(?:\s|\n|<)"),
    "email": re.compile(r"[0-9a-zA-Z.]+@[a-zA-Z.]+\.(?:nl|com|org)"),
    "username": re.compile(r"[a-z]+\d{3}"),
    "password": re.compile(r"(?<=WACHTWOORD:\s).+"),  # works only in sample.txt
}


def extract(text: str, **patterns) -> dict:
    """Extract patterns from text."""
    return {key: pattern.findall(text) for key, pattern in patterns.items()}


def write_to_bulkformulier(names, emails, usernames) -> None:
    """Writes data to new workbook using template_mutatie_medewerker.xlsx."""
    wb = load_workbook(filename='template_mutatie_medewerker.xlsx')
    sheet = wb.worksheets[0]
    for index, name in enumerate(names, 2):
        sheet.cell(index, 3).value = name
        sheet.cell(index, 2).value = emails[index - 2]
        sheet.cell(index, 1).value = usernames[index - 2]
    wb.save("mutatie_medewerker.xlsx")


def write_to_gebruikersgegevens(names, usernames, passwords, emails):
    """Writes data to new workbook.

    Using template_gebruikergegevens_instroomgroep_14020.xlsx
    """
    wb = load_workbook(filename='template_gebruikergegevens_instroomgroep_14020.xlsx')
    sheet = wb.worksheets[0]
    row_n = 1
    for i, name in enumerate(names):
        sheet.cell(row_n := row_n + 1, 1).value = "DISPLAYNAAM: " + name[0] + ", " + name[1]
        sheet.cell(row_n := row_n + 1, 1).value = "ACCOUNTNAAM: " + usernames[i]
        sheet.cell(row_n, 4).value = "ADW Gebruikernaam"
        sheet.cell(row_n := row_n + 1, 1).value = "WACHTWOORD: " + passwords[i]
        sheet.cell(row_n, 4).value = "ADW Wachtwoord"
        sheet.cell(row_n := row_n + 1, 1).value = "EMAIL: " + emails[i]
        row_n += 2
    wb.save("gebruikergegevens_instroomgroep_14020.xlsx")


def extract_write(text: str) -> None:
    """Wraps functions: extract and write_to_bulkformulier."""
    extracted_data = extract(
        text,
        last_first_names=patterns["last_first_name"],
        emails=patterns["email"],
        usernames=patterns["username"],
        passwords=patterns["password"]
    )
    full_names = [f"{first_name} {last_name}" for last_name, first_name in extracted_data["last_first_names"]]
    write_to_gebruikersgegevens(
        extracted_data["last_first_names"],
        extracted_data["usernames"],
        extracted_data["passwords"],
        extracted_data["emails"],
    )
    write_to_bulkformulier(
        full_names,
        extracted_data["emails"],
        extracted_data["usernames"],
    )


def extract_write_clear(txt_box) -> None:
    extract_write(txt_box.get("1.0", tk.END))
    txt_box.delete("1.0", tk.END)


def main():
    window = tk.Tk()

    # Box to past data in
    lbl_title = tk.Label(text="Extract contents of text box to excel")
    lbl_title.pack()
    txt_box = tk.Text()
    txt_box.pack()

    # Extract data button that starts extracting process
    btn_extract = tk.Button(
        text="Extract",
        command=lambda: extract_write_clear(txt_box)
    )
    btn_extract.pack()
    window.mainloop()


if __name__ == "__main__":
    main()

